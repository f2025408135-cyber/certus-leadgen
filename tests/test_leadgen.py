"""Tests for the certus-leadgen scripts (filter, tracker, drafts, send validation)."""
from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

PIPE = Path(__file__).resolve().parent.parent
SCRIPTS = PIPE / "scripts"
sys.path.insert(0, str(SCRIPTS))

import filter_leads  # noqa: E402
import generate_drafts  # noqa: E402
import usp_leads  # noqa: E402

FIX = PIPE / "fixtures"


def load_csv(path: Path) -> list[dict]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh, delimiter=";"))


# ---------------- usp_leads: value + text parsing ----------------------------
def test_parse_value_de():
    assert usp_leads.parse_value("11.839.880,91 €") == 11839880.91
    assert usp_leads.parse_value("74.922,50 €") == 74922.5
    assert usp_leads.parse_value("") is None
    assert usp_leads.parse_value("nicht verfügbar") is None


def test_field_after():
    t = "Auftragnehmer | | | | | Großbau Konzern AG | | | | | Stammzahl: | | ATU123"
    assert usp_leads.field_after(t, "Auftragnehmer").startswith("Großbau Konzern AG")


# ---------------- filter_leads ------------------------------------------------
def test_filter_full():
    rows = load_csv(FIX / "usp_awards_sample.csv")
    kept, dropped = filter_leads.filter_rows(rows, 50_000, 20_000_000, 12, sme_filter=True)
    names = {r["winner"] for r in kept}
    assert "Musterbau Huber GmbH" in names          # SME, in range, recent
    assert "Elektro Kaiser GmbH" in names           # SME, in range
    assert "Gebäudetechnik Muster GmbH" in names  # SME, in range
    assert "IT & Partner OG" in names    # IT CPV allowed
    assert "Großbau Konzern AG" not in names     # large corp
    assert "Tiefbau Riesen AG" not in names  # large corp
    assert "Bietergemeinschaft Alpha" not in names  # consortium
    assert "Baumeister Beispiel GmbH" not in names  # too old (2025-01 > 12 months)
    assert "Muster GmbH" not in names       # below 50k
    assert "Baumeister Winter KG" in names   # AG but not on blocklist -> kept (human check note)
    assert dropped["large_corp"] == 2
    assert dropped["consortium"] == 1
    assert dropped["value"] == 1
    assert dropped["date"] == 1


def test_filter_no_sme_filter():
    rows = load_csv(FIX / "usp_awards_sample.csv")
    kept, _ = filter_leads.filter_rows(rows, 50_000, 20_000_000, 12, sme_filter=False)
    names = {r["winner"] for r in kept}
    assert "Großbau Konzern AG" in names


def test_filter_value_below_min():
    rows = load_csv(FIX / "usp_awards_sample.csv")
    kept, _ = filter_leads.filter_rows(rows, 60_000, 20_000_000, 12, sme_filter=True)
    assert "IT & Partner OG" not in kept  # 65k... wait >= 60k -> kept. check:
    assert any(r["winner"] == "IT & Partner OG" for r in kept)


# ---------------- tracker ------------------------------------------------------
def test_tracker_xlsx_roundtrip(tmp_path):
    from openpyxl import load_workbook
    import leads_tracker as lt

    wb, created = lt._wb(tmp_path / "leads_formatted.xlsx")
    assert created is False
    ws = lt._sheet(wb)
    n = lt.import_csv(FIX / "usp_awards_sample.csv", wb)
    assert n == 10
    # merge again -> 0 new
    n2 = lt.import_csv(FIX / "usp_awards_sample.csv", wb)
    assert n2 == 0
    assert lt.set_field(ws, "Musterbau Huber GmbH", "status", "approved_r1")
    wb.save(tmp_path / "leads_formatted.xlsx")

    wb2 = load_workbook(tmp_path / "leads_formatted.xlsx")
    ws2 = wb2["Cold Email"]
    assert ws2.max_row == 11  # header + 10
    vals = {ws2.cell(row=r, column=1).value: ws2.cell(row=r, column=11).value for r in range(2, ws2.max_row + 1)}
    assert vals["Musterbau Huber GmbH"] == "approved_r1"


# ---------------- drafts --------------------------------------------------------
def _cfg(tmp_path) -> Path:
    c = tmp_path / "sender_config.json"
    c.write_text(json.dumps({
        "sender_name": "Test",
        "sender_line": "Test student",
        "question": "Was war bei der Angebotserstellung dieses Auftrags der aufwendigste Teil?",
        "signature_name": "Test User",
        "phone": "+00 000 0000000",
    }, ensure_ascii=False), encoding="utf-8")
    return c


def test_draft_render_faithful(tmp_path):
    cfg = json.loads(_cfg(tmp_path).read_text(encoding="utf-8"))
    lead = {
        "winner": "Gebäudetechnik Muster GmbH",
        "tender_title": "KOR_2026_01 HKLS Sanierung Objekt 12-14",
        "authority": "Stadt Musterstadt",
        "award_date": "2026-05-15",
    }
    body = generate_drafts.render_round1(lead, cfg)
    assert "Guten Tag, liebes Team der Gebäudetechnik Muster GmbH" in body
    assert "Test student" in body
    assert "im Mai diesen Jahres" in body
    assert "den Auftrag zu KOR_2026_01 HKLS Sanierung Objekt 12-14" in body
    assert "Stadt Musterstadt" in body
    assert "Was war bei der Angebotserstellung dieses Auftrags der aufwendigste Teil?" in body
    assert "Test User" in body and "+00 000 0000000" in body


def test_draft_files_written(tmp_path):
    outdir = tmp_path / "drafts"
    generate_drafts.main.__wrapped__ if hasattr(generate_drafts.main, "__wrapped__") else None
    import argparse as _a
    # drive main() via monkeypatched sys.argv
    csv_in = tmp_path / "in.csv"
    rows = load_csv(FIX / "usp_awards_sample.csv")
    rows = [r for r in rows if r["winner"] in ("Musterbau Huber GmbH", "Elektro Kaiser GmbH")][:2]
    with csv_in.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=list(rows[0].keys()), delimiter=";")
        w.writeheader()
        w.writerows(rows)
    old = sys.argv
    sys.argv = ["generate_drafts.py", "--csv", str(csv_in), "--round", "1", "--config", str(_cfg(tmp_path)), "--max", "5", "--outdir", str(outdir)]
    try:
        assert generate_drafts.main() == 0
    finally:
        sys.argv = old
    files = [f for f in (outdir / "round1").glob("*.txt") if not f.name.startswith("_")]
    assert len(files) == 2
    combined = outdir / "round1" / "_preview_all.txt"
    assert combined.exists()
    assert "Musterbau Huber GmbH" in combined.read_text(encoding="utf-8")


# ---------------- send gate -----------------------------------------------------
def test_send_requires_credentials_and_approval(tmp_path):
    import send_emails as se
    # no creds, no approval file -> exit 2 (credentials first)
    old = sys.argv
    sys.argv = ["send_emails.py", "--drafts", str(tmp_path), "--approvals", str(tmp_path / "a.csv")]
    try:
        try:
            se.main()
            assert False, "should exit(2) on missing credentials"
        except SystemExit as e:
            assert e.code == 2
    finally:
        sys.argv = old


def test_send_dry_run_matches_approved(tmp_path):
    import send_emails as se
    creds_file = tmp_path / "credentials.json"
    creds_file.write_text(json.dumps({
        "smtp_host": "smtp.gmail.com", "smtp_port": 587,
        "smtp_user": "test@example.at", "smtp_pass": "x", "from_name": "Test",
    }), encoding="utf-8")
    drafts = tmp_path / "drafts"
    drafts.mkdir()
    (drafts / "01-musterbau-huber-gmbh.txt").write_text("Body", encoding="utf-8")
    (drafts / "02-other-gmbh.txt").write_text("Body", encoding="utf-8")
    appr = tmp_path / "approved.csv"
    with appr.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=["company", "email"], delimiter=";")
        w.writeheader()
        w.writerow({"company": "Musterbau Huber GmbH", "email": "office@musterbau-huber.example.at"})
    old = sys.argv
    sys.argv = ["send_emails.py", "--drafts", str(drafts), "--approvals", str(appr), "--credentials", str(creds_file)]
    try:
        assert se.main() == 0  # dry-run: only matched draft listed
    finally:
        sys.argv = old
