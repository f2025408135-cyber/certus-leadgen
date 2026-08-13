"""Lead tracker — leads_formatted.xlsx with sheet 'Cold Email' + per-round status.

Columns (per the founder brief): company, website, email, contact_person,
tender_title, authority, award_date, value_eur, industry, source_url,
plus status tracking per email round: status, round1_status, round1_sent,
round1_reply, round2_status, round2_sent.

Usage:
  python leads_tracker.py import <filtered_or_researched.csv>   # merge new leads
  python leads_tracker.py show                                  # summary
  python leads_tracker.py set <company> <field> <value>         # update a field
  python leads_tracker.py approve <company> [--round 1]         # mark approved for sending
"""
from __future__ import annotations

import argparse
import csv
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DEFAULT_PATH = Path("marketing/leads/leads_formatted.xlsx")
SHEET = "Cold Email"

COLUMNS = [
    "company", "website", "email", "contact_person", "tender_title", "authority",
    "award_date", "value_eur", "industry", "source_url", "status",
    "round1_status", "round1_sent", "round1_reply", "round2_status", "round2_sent", "notes",
]

STATUS_NEW = "new"
HEADER_FILL = PatternFill("solid", fgColor="C2410C")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _wb(path: Path) -> tuple[Workbook, bool]:
    if path.exists():
        return load_workbook(path), True
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(COLUMNS)
    for c in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    return wb, False


def _sheet(wb: Workbook):
    if SHEET in wb.sheetnames:
        return wb[SHEET]
    ws = wb.create_sheet(SHEET)
    ws.append(COLUMNS)
    return ws


def _rows(ws) -> dict[str, dict]:
    out = {}
    for r in range(2, ws.max_row + 1):
        vals = {COLUMNS[i]: (ws.cell(row=r, column=i + 1).value or "") for i in range(len(COLUMNS))}
        key = vals["company"].strip().lower()
        if key:
            out[key] = {"row": r, **vals}
    return out


def import_csv(path: Path, wb: Workbook) -> int:
    ws = _sheet(wb)
    existing = _rows(ws)
    with path.open(encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.DictReader(fh, delimiter=";"))
    added = 0
    for r in rows:
        company = (r.get("winner") or r.get("company") or "").strip()
        if not company:
            continue
        key = company.lower()
        row_vals = {
            "company": company,
            "website": r.get("website", ""),
            "email": r.get("email", ""),
            "contact_person": r.get("contact_person", ""),
            "tender_title": r.get("tender_title", ""),
            "authority": r.get("authority", ""),
            "award_date": r.get("award_date", ""),
            "value_eur": r.get("value_eur_parsed") or r.get("value_eur", ""),
            "industry": r.get("industry", ""),
            "source_url": r.get("source_url", ""),
            "status": r.get("status", STATUS_NEW),
            "notes": r.get("notes", ""),
        }
        if key in existing:
            # merge: fill only empty cells (preserve status history)
            row = existing[key]["row"]
            for i, col in enumerate(COLUMNS):
                if not (ws.cell(row=row, column=i + 1).value or ""):
                    ws.cell(row=row, column=i + 1).value = row_vals.get(col, "")
        else:
            ws.append([row_vals.get(c, "") for c in COLUMNS])
            added += 1
    return added


def set_field(ws, company: str, field: str, value: str) -> bool:
    if field not in COLUMNS:
        raise ValueError(f"unknown field {field}")
    idx = COLUMNS.index(field)
    for r in range(2, ws.max_row + 1):
        if (ws.cell(row=r, column=1).value or "").strip().lower() == company.lower():
            ws.cell(row=r, column=idx + 1).value = value
            return True
    return False


def show(ws) -> None:
    rows = _rows(ws)
    if not rows:
        print("[tracker] empty")
        return
    by_status: dict[str, int] = {}
    for v in rows.values():
        st = v["status"] or STATUS_NEW
        by_status[st] = by_status.get(st, 0) + 1
    print(f"[tracker] {len(rows)} leads; status: {by_status}")
    for v in list(rows.values())[:12]:
        print(f"  {v['company'][:38]:40} {v['value_eur']!s:>14} {v['award_date'][:10]:10} {v['status']:12} {v['round1_status']}")


def main() -> int:
    p = argparse.ArgumentParser(prog="leads_tracker.py")
    sub = p.add_subparsers(dest="cmd", required=True)
    imp = sub.add_parser("import")
    imp.add_argument("csv_file")
    showp = sub.add_parser("show")
    setp = sub.add_parser("set")
    setp.add_argument("company")
    setp.add_argument("field")
    setp.add_argument("value")
    appr = sub.add_parser("approve")
    appr.add_argument("company")
    appr.add_argument("--round", default="1")
    args = p.parse_args()

    wb, _ = _wb(DEFAULT_PATH)
    ws = _sheet(wb)
    if args.cmd == "import":
        n = import_csv(Path(args.csv_file), wb)
        print(f"[tracker] imported {n} new leads -> {DEFAULT_PATH}")
    elif args.cmd == "show":
        show(ws)
    elif args.cmd == "set":
        ok = set_field(ws, args.company, args.field, args.value)
        print(f"[tracker] {'updated' if ok else 'not found'}: {args.company} {args.field}={args.value}")
    elif args.cmd == "approve":
        field = f"round{args.round}_status"
        ok = set_field(ws, args.company, "status", f"approved_r{args.round}") and set_field(ws, args.company, field, "approved")
        print(f"[tracker] {'approved for round ' + args.round if ok else 'not found'}: {args.company}")
    wb.save(DEFAULT_PATH)
    return 0


if __name__ == "__main__":
    sys.exit(main())
